#coding=gbk
import ftplib
import os
import time
import shutil
from pandas import Series, DataFrame
import pandas as pd
import numpy as np
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import subprocess
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import re
import itchat

#�ļ������ļ�����
target_file_name = 'AAAAAAA_' #����
target_file_type = '.txt.Z'

#ftp�ļ����Ŀ¼
ftpfilepath = '/��ʱ�嵥'

#�����ļ����Ŀ¼
saverootpath = 'D:\\'

#winrar��װĿ¼
winrarpath = 'D:\\winrar\\winrar'

#Excel��װĿ¼
officepath = 'D:\\Microsoft Office\\Office12\\excel.exe'

def get_file_name(ftpnamelist):
	'''
	��ȡ���ص��ļ�����
	'''
	
	all_target_file_name = []
	
	for item in ftpnamelist:
		if item[:len(target_file_name)] == target_file_name:
			all_target_file_name.append(int(item[len(target_file_name):len(target_file_name) + 8]))

	newestdate = max(all_target_file_name)

	print('��ǰFTP�����嵥Ϊ�� ' + str(newestdate))
	
	print('\n����س���ȡ�����嵥�������������ڣ���ȡ���������嵥:')
	
	while True:
		get = input()

		if get == '':
			a = str(newestdate)
		elif int(get) in all_target_file_name:
			a = get
		else:
			print('����������ڸ�ʽ����򲻴��ڸ������嵥�����������룺')
			continue
			
		downloadfilename = target_file_name + a + target_file_type
		break
	
	return downloadfilename


def down_file_from_ftp():
	'''
	��¼FTP�����ļ�����Ҫ��pythonĿ¼�е�ftplib.py��encoding = "lantin-1"�޸�Ϊencoding = "gbk"
	'''
	
	#FTP��¼��Ϣ
	host = 'AAA.AA.AAA.153' #����
	username = 'AAAAAA' #����
	password = 'AAAAAAAAA' #����

	#��¼FTP
	ftp = ftplib.FTP(host)
	ftp.login(username, password)
	print('�ɹ���¼FTP')
	print(ftp.getwelcome())	

	#����·��
	ftp.cwd(ftpfilepath)
	current_path = ftp.pwd()
	print('\nFTP��ǰ·������Ϊ��', current_path)

	#��ȡ·���������ļ���
	allftpfilename = ftp.nlst()

	#������Ҫ���ص��ļ���
	downloadfilename = get_file_name(allftpfilename)

	#���������ļ��Ĵ��Ŀ¼
	savepath = saverootpath + downloadfilename

	#�����ļ���ָ��Ŀ¼
	print('��������"' + downloadfilename + '"  ...')
	copyfile = open(savepath, 'wb')  
	needfilename = 'RETR ' + downloadfilename 
	ftp.retrbinary(needfilename, copyfile.write)
	
	#������Ϻ��1���ٴ�ӡ��Ϣ�������ѹʱ�ļ�����ʹ��
	time.sleep(1)
	print('�ļ��������')

	#�ǳ�FTP
	ftp.quit() 
	print('\n�ǳ�FTP')
	
	#�������ص��ļ����������·��
	return downloadfilename, savepath


def unrar_the_file(rarfilename, rarpath):
	'''
	����ϵͳwinrar�����ѹ�ļ�
	'''

	#��ѹ���ص��ļ�
	print('\n���ڽ�ѹ"' + rarfilename + '"  ...')
	cmd = winrarpath + ' x -ibck "' + rarpath+ '" "' + saverootpath + '"'
	os.system(cmd)#������
	time.sleep(1)
	print('��ѹ���')
	
	unrarfilename = rarfilename.replace('.Z', '')
	unrarpath = rarpath.replace('.Z', '')
	return unrarfilename, unrarpath


def change_filename_to_english(chinesefilename, chinesepath):
	'''
	�������ļ����޸�ΪӢ���ļ���
	'''
	englishfilename = chinesefilename[len(target_file_name):]
	englishpath = chinesepath[:len(chinesepath)-len(chinesefilename)] + englishfilename
	shutil.move(chinesepath, englishpath)
	
	return englishfilename, englishpath
	

def data_statistic(englishfilename, englishpath):
	'''
	��������ͳ�ƣ������Excel����
	'''

	#�����ݵ���ΪDataFrame
	df = pd.read_table(englishpath, sep='$', encoding='gbk', low_memory = False)
	print('\n���ڽ�������ͳ��...')

	#�������и�ʽ��String�޸�Ϊdatetime
	timelist_str = pd.Series(df['�һ�����'])
	timelist_datetime = pd.Series(datetime.datetime.strptime(k, '%Y-%m-%d') for k in timelist_str)
	df['�һ�����'] = timelist_datetime

	#��������
	districtindex = ['A', 'AA', 'AAA', 'AAA', 'A', 'A', 'A', 'A', 'A', 'A', 'A', 'A', 'A', 'A'] #����
	filedatestr = englishfilename.replace('.txt', '')
	filedatedatime = datetime.datetime.strptime(filedatestr, '%Y%m%d')
	delta = datetime.timedelta(days = 7)
	groupbykeyword = 'AAAA' #����
	sumkeyword = 'AAAA' #����

	#ͳ��ȫ�����
	yearstatistic_unsort = df.groupby(groupbykeyword).agg({groupbykeyword : 'count', sumkeyword : sum})
	yearstatistic_sort = yearstatistic_unsort.reindex(districtindex,fill_value=0)

	#ͳ�Ƶ������
	monthformat = lambda x : x.month
	df_month = df[df['�һ�����'].map(monthformat) == filedatedatime.month]
	monthstatistic_unsort = df_month.groupby(groupbykeyword).agg({groupbykeyword : 'count', sumkeyword : sum})
	monthstatistic_sort = monthstatistic_unsort.reindex(districtindex,fill_value=0)

	#ͳ���������
	df_week = df[df['�һ�����'] > (filedatedatime - delta)]
	weekstatistic_unsort = df_week.groupby(groupbykeyword).agg({groupbykeyword : 'count', sumkeyword : sum})
	weekstatistic_sort = weekstatistic_unsort.reindex(districtindex,fill_value=0)

	#�ϲ�3��DataFrame���޸��������Լ����Ӻϼ���
	week_month_df = pd.merge(weekstatistic_sort, monthstatistic_sort, left_index=True, right_index=True, suffixes=('_week', '_month'))
	week_month_year_df = pd.merge(week_month_df, yearstatistic_sort, left_index=True, right_index=True,)
	week_month_year_df.columns = [['����','����','����','����','2018��','2018��'], ['AAA������', 'AAAA����Ԫ��']*3] #����
	week_month_year_df.loc['�ϼ�'] = week_month_year_df.sum()
	print('����ͳ�����')

	#���excel�ļ�
	print('\n�������Excel����...')
	excelfilepath = englishpath.replace('.txt', '.xlsx')
	week_month_year_df.to_excel(excelfilepath, encoding = 'gbk')
	print('Excel���������ϣ���ŵ�ַΪ��' + excelfilepath)
	
	return excelfilepath
	

def adjust_excel_style(excelfilepath):
	'''
	����Excel������ʽ
	'''
	#��Excel�ļ�
	wb = openpyxl.load_workbook(excelfilepath)
	sheet = wb.active

	print('\n���ڵ���Excel������ʽ...')

	#�����ͷ
	sheet.insert_rows(0)
	sheet.unmerge_cells('B1:C1')
	sheet.unmerge_cells('D1:E1')
	sheet.unmerge_cells('F1:G1')
	sheet.merge_cells('B2:C2')
	sheet.merge_cells('D2:E2')
	sheet.merge_cells('F2:G2')
	sheet['A1'] = 'AAAA�������' #����
	sheet.merge_cells('A1:G1')
	sheet['A2'] = sheet['A4'].value
	sheet.merge_cells('A2:A3')
	sheet.delete_rows(4)
	
	#����ע
	

	#�����и�
	sheet.row_dimensions[1].height = 40
	sheet.row_dimensions[2].height = 17.4
	sheet.row_dimensions[3].height = 34.8
	sheet.row_dimensions[18].height = 17.4

	for i in range(4,18):
		sheet.row_dimensions[i].height = 15


	#�����п�
	sheet.column_dimensions['A'].width = 10
	for i in 'BCDEFG':
		sheet.column_dimensions[i].width = 15


	#���õ�Ԫ�����塢���롢�߿򡢱�����ɫ��ʽ
	titleFont = Font(name = '΢���ź�', size = 20, bold = True, italic = False)
	headFont = Font(name = '΢���ź�', size = 12, bold = True, italic = False)
	cellFont = Font(name = '΢���ź�', size = 11, bold = False, italic = False)

	titleAlignment = Alignment(horizontal = 'center', vertical = 'bottom')
	cellAlignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text= True)

	cellBorder = Border(left = Side(border_style = 'thin', color = 'FF000000'),
					right = Side(border_style = 'thin', color = 'FF000000'),
					top = Side(border_style = 'thin', color = 'FF000000'),
					bottom = Side(border_style = 'thin', color = 'FF000000'))

	cellPatternFill = PatternFill(fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')

	#����ʽӦ���ڱ���
	sheet['A1'].font = titleFont
	sheet['A1'].alignment = titleAlignment

	#����ʽӦ���ڱ�ͷ
	for RowofCell in sheet['A2':'G3']:
		for Cell in RowofCell:
			Cell.font = headFont
			Cell.alignment = cellAlignment
			Cell.border = cellBorder
			Cell.fill = cellPatternFill

	#����ʽӦ���ڵ�Ԫ��
	for RowofCell in sheet['A4':'G17']:
		for Cell in RowofCell:
			Cell.font = cellFont
			Cell.alignment = cellAlignment
			Cell.border = cellBorder

	#����ʽӦ���ںϼ���
	for RowofCell in sheet['A18':'G18']:
		for Cell in RowofCell:
			Cell.font = headFont
			Cell.alignment = cellAlignment
			Cell.border = cellBorder
			Cell.fill = cellPatternFill
        
	print('Excel������ʽ�������')

	#����Excel���
	wb.save(excelfilepath)
	
	return True
	
	
def open_excel(state):
	if state:
		print('\n���ڴ�Excel����...')
		subprocess.Popen([officepath, excelfilepath]).wait()
		print('Excel�����ѹر�')
		

def add_input_to_list(emaillist, to_list, terminal_staff, Falseemail):
	for item in emaillist:
		if '@' in item:
			to_list.append(item)
		elif item in terminal_staff:
			to_list.append(terminal_staff[item])
		else:
			Falseemail.append(item)		
		
		
def add_email_to_list(inputstr, to_list, terminal_staff):
	'''
	��������ַ���ת��Ϊemail�б�
	'''
			
	email_list = inputstr.replace(' ', '').split(',')
	Falseemail = []
			
	add_input_to_list(email_list, to_list, terminal_staff, Falseemail)

	while len(Falseemail):
				
		if len(to_list) > 1:
			print('\n���µ��������ַ�Ѽ��뷢�����б�')
			for item in to_list[1:]:
				print(item)
						
		print('�������' + ', '.join(Falseemail) + '������ָ����Ա����������ϸ�ʼ���ַ��')
		print('����no������ӣ�')
		secondinput = input()
				
		if secondinput == 'no' and to_list == 1:
			print('�����͵����ʼ�')
			return False
		elif secondinput == 'no':
			break
		else:
			email_list2 = secondinput.replace(' ', '').split(',')
			Falseemail = []
			add_input_to_list(email_list2, to_list, terminal_staff, Falseemail)

		
def send_email(excelfilepath):
	'''
	��Excel�����͵����ʼ���ָ����Ա
	'''

	#���ñ���
	my_email_address = 'AAAA@AAAA.cn' #����
	my_pas = 'AAAAAAA' #����
	to_list = [my_email_address,]
	terminal_staff = {'AAA' : 'AAAA@AAAA.cn',
					  'AAAAA' : 'AAAAAA@AAAA.cn',} #����

	
	print('\n�Ƿ�Excel�����͵����ʼ���ָ����Ա��')
	print('������������OA�˻�������ƴ����OA��ƴ������ָ����Ա�������Ͷ����Ա��Ӣ�Ķ��ŷֿ�������no�����ͣ�')
	sendornot = input()
	
	#�Ƿ����ʼ�
	if sendornot == 'no':
		print('�����͵����ʼ�')
		return False
	else:
		add_email_to_list(sendornot, to_list, terminal_staff)
		

	#�����ʼ����⼰��������
	msg = MIMEMultipart()
	msg['Subject'] = 'AAAAҵ���������' + re.sub('\D', '', excelfilepath) #����
	puretext = MIMEText('������')
	msg.attach(puretext)

	#�����ʼ���������
	xlsxpart = MIMEApplication(open(excelfilepath, 'rb').read())
	xlsxpart.add_header('Content-Disposition', 'attachment', filename = excelfilepath)
	msg.attach(xlsxpart)

	#���͵����ʼ�
	print('\n���ڷ��͵����ʼ�...')
	smtpObj = smtplib.SMTP_SSL('smtp.chinatelecom.cn', 465)
	smtpObj.ehlo()
	smtpObj.login(my_email_address, my_pas)
	smtpObj.sendmail(my_email_address, to_list, msg.as_string())
	smtpObj.quit()
	
	print('\n�ѽ�Excel���������������䣺')
	for item in to_list:
		print(item)
		

def send_excel_by_wechat(excelfilepath):
	'''
	ͨ��΢�Ž�Excel������ָ����Ա
	'''

	print('\n�Ƿ�Excel����ͨ��΢�ŷ��͸�ָ����Ա��')
	print('����yes��no��')
	sendornot = input()
	
	#�Ƿ�ͨ��΢�ŷ���
	if sendornot == 'yes':
		print('\n���ڵ�¼΢��...')
		itchat.auto_login(hotReload = True)
		print('��¼΢�ųɹ�')
	
		lmy = itchat.search_friends(nickName = '������')
		lmy_username = lmy[0]['UserName']
	
		print('\n����ʹ��΢�ŷ���Excel����')
		itchat.send_file(excelfilepath, lmy_username)
		itchat.send_msg('����Ϊ' + re.sub('\D', '', excelfilepath) + 'AAAҵ���������������ա�', lmy_username)#����
		print('�������')
	
		itchat.logout()
		print('�ǳ�΢��')
	
	
def remove_file(rarfilepath, txtfilepath, excelfilepath):
	'''
	ɾ���ļ�
	'''
	print('\n�Ƿ���Ҫɾ��Դ�ļ�������yɾ��������������ɾ����')
	delornot = input()
	
	if delornot == 'y':
		os.remove(rarfilepath)
		os.remove(txtfilepath)
		print('\n��ɾ�������ļ���')
		print(rarfilepath)
		print(txtfilepath)
		
		print('\n�Ƿ���Ҫɾ��Excel�����ļ�������yɾ��������������ɾ����')
		exceldelornot = input()
		
		if exceldelornot == 'y':
			os.remove(excelfilepath)
			print('\n��ɾ�������ļ���')
			print(excelfilepath)
	
	print('\n��л����ʹ��')


#��¼FTP���������ļ����������ļ���������·��
downloadfilename, savepath = down_file_from_ftp()

#��ѹ�ļ�
unrarfilename, unrarpath = unrar_the_file(downloadfilename, savepath)

#�޸�������ΪӢ����
englishfilename, englishpath = change_filename_to_english(unrarfilename, unrarpath)

#ͳ�����ݲ����Excel����
excelfilepath = data_statistic(englishfilename, englishpath)

#����Excel������ʽ
state = adjust_excel_style(excelfilepath)

#��Excel����
open_excel(state)
	
#��Excel�����͵����ʼ���ָ����Ա
send_email(excelfilepath)

#��Excel����ͨ��΢�ŷ��͸�ָ����Ա
send_excel_by_wechat(excelfilepath)

#ɾ���ļ�
remove_file(savepath, englishpath, excelfilepath)


